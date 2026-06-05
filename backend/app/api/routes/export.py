import io
from datetime import date

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from sqlalchemy.orm import Session, joinedload

from app.api.deps import get_db_session
from app.models.athlete import Athlete
from app.models.attendance import Attendance
from app.models.result import Result
from app.models.training_session import TrainingSession

router = APIRouter(prefix="/export", tags=["Export"])


@router.get("", summary="Export all data to Excel")
def export_excel(db: Session = Depends(get_db_session)) -> StreamingResponse:
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Athletes")
    ws.append(["ID", "Name", "Date of Birth", "Sex", "Height (cm)", "Weight (kg)",
               "Shirt Size", "Short Size", "Shoe Size", "Active", "Notes"])
    for a in db.query(Athlete).order_by(Athlete.name).all():
        ws.append([a.id, a.name, a.date_of_birth, a.sex, a.height, a.weight,
                   a.shirt_size, a.short_size, a.shoe_size,
                   "Yes" if a.is_active else "No", a.notes or ""])

    ws = wb.create_sheet("Training Sessions")
    ws.append(["ID", "Date", "Title", "Notes"])
    for s in db.query(TrainingSession).order_by(TrainingSession.date.desc()).all():
        ws.append([s.id, s.date, s.title or "", s.notes or ""])

    ws = wb.create_sheet("Attendance")
    ws.append(["ID", "Athlete", "Session Date", "Status", "Notes"])
    for r in db.query(Attendance).options(joinedload(Attendance.athlete), joinedload(Attendance.session)).all():
        ws.append([r.id, r.athlete.name if r.athlete else "",
                   r.session.date if r.session else "", r.status, r.notes or ""])

    ws = wb.create_sheet("Results")
    ws.append(["ID", "Athlete", "Session Date", "Event", "Value", "Unit", "Result Date", "Notes"])
    for r in db.query(Result).options(joinedload(Result.athlete), joinedload(Result.session)).all():
        ws.append([r.id, r.athlete.name if r.athlete else "",
                   r.session.date if r.session else "",
                   r.event_name, r.value, r.unit, r.result_date, r.notes or ""])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"training_logs_{date.today()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )